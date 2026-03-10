from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
import openpyxl
import os
from datetime import datetime
import firebase_admin
from firebase_admin import credentials, firestore, initialize_app
import json
from io import BytesIO
import traceback
import re
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max para archivos grandes

# Inicializar Firebase con las credenciales del archivo bd.txt
firebase_config = {
    "type": "service_account",
    "project_id": "proyectosalamanca-e1719",
    "private_key_id": "",
    "private_key": "",
    "client_email": "",
    "client_id": "",
    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
    "token_uri": "https://oauth2.googleapis.com/token",
    "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs"
}

# Inicializar Firebase (solo si no está inicializado)
try:
    firebase_admin.get_app()
except ValueError:
    # Si no hay credenciales de servicio, usar configuración básica
    cred = credentials.Certificate(firebase_config) if os.path.exists('firebase-credentials.json') else None
    if cred:
        initialize_app(cred)
    else:
        # Inicializar sin credenciales para usar desde el cliente
        pass

# Obtener referencia a Firestore (si está disponible)
try:
    db = firestore.client()
except:
    db = None

MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

def limpiar_texto(valor):
    """Limpia un valor de texto quitando espacios extra y caracteres invisibles"""
    if valor is None:
        return ''
    texto = str(valor).strip()
    # Quitar caracteres invisibles y espacios múltiples
    texto = re.sub(r'\s+', ' ', texto)
    return texto

def coincide_encabezado(texto_celda, palabras_clave):
    """Verifica si el texto de una celda coincide con alguna de las palabras clave (flexible)"""
    if not texto_celda:
        return False
    texto = limpiar_texto(texto_celda).lower()
    # Quitar tildes para comparación flexible
    reemplazos = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n'}
    for orig, reempl in reemplazos.items():
        texto = texto.replace(orig, reempl)
    
    for palabra in palabras_clave:
        if palabra.lower() in texto:
            return True
    return False

def extraer_datos_excel(file_stream):
    """Extrae datos de las hojas de meses del archivo Excel desde un stream"""
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=False)
    except Exception as e1:
        # Intentar sin data_only
        try:
            file_stream.seek(0)
            wb = openpyxl.load_workbook(file_stream, data_only=False, read_only=False)
        except Exception as e2:
            raise Exception(f'No se pudo abrir el archivo Excel. Verifica que sea un archivo .xlsx o .xlsm válido. Error: {str(e1)}')
    
    datos_por_mes = {}
    
    # Mapeo flexible de nombres de hojas a meses
    mapeo_hojas = {}
    for nombre_hoja in wb.sheetnames:
        hoja_limpia = limpiar_texto(nombre_hoja).lower()
        # Quitar tildes
        reemplazos = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u'}
        for orig, reempl in reemplazos.items():
            hoja_limpia = hoja_limpia.replace(orig, reempl)
        
        for mes in MESES:
            mes_limpio = mes.lower()
            for orig, reempl in reemplazos.items():
                mes_limpio = mes_limpio.replace(orig, reempl)
            
            if mes_limpio in hoja_limpia or hoja_limpia in mes_limpio:
                mapeo_hojas[nombre_hoja] = mes
                break
            # También verificar abreviaciones (Ene, Feb, Mar, etc.)
            abrev = mes_limpio[:3]
            if hoja_limpia.startswith(abrev) or hoja_limpia == abrev:
                mapeo_hojas[nombre_hoja] = mes
                break
    
    for nombre_hoja, mes in mapeo_hojas.items():
        try:
            ws = wb[nombre_hoja]
            datos_mes = []
            
            # Buscar la fila de encabezados - buscar más amplio
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30), start=1):
                for cell in row:
                    if cell.value and coincide_encabezado(str(cell.value), ['cedula', 'cédula', 'cedul', 'c.c', 'documento', 'identificacion']):
                        header_row = row_idx
                        break
                if header_row:
                    break
            
            if not header_row:
                print(f'No se encontró encabezado en la hoja {nombre_hoja}, saltando...')
                continue
            
            # Obtener índices de columnas
            headers = {}
            for cell in ws[header_row]:
                if cell.value:
                    val = limpiar_texto(cell.value)
                    headers[val] = cell.column
            
            # Buscar columnas con nombres alternativos (más flexible)
            cedula_col = None
            nombres_col = None
            ingreso_col = None
            cargo_col = None
            distrito_col = None
            dias_pendientes_col = None
            
            for key, col in headers.items():
                key_lower = key.lower()
                # Quitar tildes para comparación
                key_sin_tildes = key_lower
                reemplazos_tildes = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n'}
                for orig, reempl in reemplazos_tildes.items():
                    key_sin_tildes = key_sin_tildes.replace(orig, reempl)
                
                if any(x in key_sin_tildes for x in ['cedula', 'c.c', 'documento', 'identificacion']):
                    cedula_col = col
                elif any(x in key_sin_tildes for x in ['apellido', 'nombre', 'empleado', 'trabajador', 'funcionario']):
                    nombres_col = col
                elif any(x in key_sin_tildes for x in ['ingreso', 'fecha_ingreso', 'vinculacion', 'contratacion']):
                    ingreso_col = col
                elif any(x in key_sin_tildes for x in ['cargo', 'puesto', 'posicion', 'funcion']):
                    cargo_col = col
                elif any(x in key_sin_tildes for x in ['distrito', 'sede', 'ubicacion', 'zona', 'regional']):
                    distrito_col = col
                elif any(x in key_sin_tildes for x in ['dias', 'pendiente']):
                    if 'pendiente' in key_sin_tildes or 'dias' in key_sin_tildes:
                        dias_pendientes_col = col
            
            # Leer datos
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                try:
                    if not row or len(row) == 0:
                        continue
                    
                    cedula_idx = (cedula_col - 1) if cedula_col else 0
                    nombres_idx = (nombres_col - 1) if nombres_col else 1
                    ingreso_idx = (ingreso_col - 1) if ingreso_col else 2
                    cargo_idx = (cargo_col - 1) if cargo_col else 3
                    distrito_idx = (distrito_col - 1) if distrito_col else 4
                    dias_pendientes_idx = (dias_pendientes_col - 1) if dias_pendientes_col else 5
                    
                    # Verificar que la fila tenga datos en la columna de cédula
                    cedula_val = row[cedula_idx] if cedula_idx < len(row) else None
                    if not cedula_val:
                        continue
                    
                    # Limpiar cédula (puede venir como número o texto)
                    cedula_limpia = str(cedula_val).strip()
                    # Quitar decimales si viene como float (ej: 12345.0 -> 12345)
                    try:
                        cedula_num = float(cedula_limpia)
                        if cedula_num == int(cedula_num):
                            cedula_limpia = str(int(cedula_num))
                    except:
                        pass
                    
                    if not cedula_limpia or cedula_limpia == 'None' or cedula_limpia == '':
                        continue
                    
                    # Formatear fecha correctamente
                    f_ingreso = row[ingreso_idx] if ingreso_idx < len(row) else ''
                    if isinstance(f_ingreso, datetime):
                        f_ingreso = f_ingreso.strftime('%d/%m/%Y')
                    elif f_ingreso:
                        f_ingreso = str(f_ingreso).strip()
                        # Intentar parsear formatos de fecha comunes
                        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y']:
                            try:
                                f_ingreso = datetime.strptime(f_ingreso, fmt).strftime('%d/%m/%Y')
                                break
                            except:
                                continue
                    else:
                        f_ingreso = ''
                    
                    # Obtener días pendientes y redondear
                    dias_pendientes = row[dias_pendientes_idx] if dias_pendientes_idx < len(row) else 0
                    if dias_pendientes is None:
                        dias_pendientes = 0
                    else:
                        try:
                            dias_pendientes = round(float(dias_pendientes))
                        except:
                            dias_pendientes = 0
                    
                    # Obtener nombres limpio
                    nombres_val = row[nombres_idx] if nombres_idx < len(row) else ''
                    nombres_val = limpiar_texto(nombres_val) if nombres_val else ''
                    
                    # Obtener cargo limpio
                    cargo_val = row[cargo_idx] if cargo_idx < len(row) else ''
                    cargo_val = limpiar_texto(cargo_val) if cargo_val else ''
                    
                    # Obtener distrito limpio
                    distrito_val = row[distrito_idx] if distrito_idx < len(row) else ''
                    distrito_val = limpiar_texto(distrito_val) if distrito_val else ''
                    
                    registro = {
                        'cedula': cedula_limpia,
                        'nombres': nombres_val,
                        'f_ingreso': f_ingreso,
                        'cargo': cargo_val,
                        'distrito': distrito_val,
                        'dias_pendientes': dias_pendientes
                    }
                    
                    datos_mes.append(registro)
                    
                except Exception as error_fila:
                    print(f'Error procesando fila en hoja {nombre_hoja}: {str(error_fila)}')
                    continue
            
            if datos_mes:
                datos_por_mes[mes] = datos_mes
                print(f'Hoja {nombre_hoja} ({mes}): {len(datos_mes)} registros extraídos')
                
        except Exception as error_hoja:
            print(f'Error procesando hoja {nombre_hoja}: {str(error_hoja)}')
            continue
    
    wb.close()
    return datos_por_mes

def organizar_por_distrito(datos_por_mes):
    """Organiza los datos por distrito"""
    datos_por_distrito = {}
    
    for mes, registros in datos_por_mes.items():
        for registro in registros:
            distrito = registro['distrito'] or 'SIN DISTRITO'
            
            if distrito not in datos_por_distrito:
                datos_por_distrito[distrito] = {}
            
            if mes not in datos_por_distrito[distrito]:
                datos_por_distrito[distrito][mes] = []
            
            datos_por_distrito[distrito][mes].append(registro)
    
    return datos_por_distrito

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/firebase-config', methods=['GET'])
def get_firebase_config():
    """Endpoint para obtener la configuración de Firebase"""
    config = {
        "apiKey": os.getenv("FIREBASE_API_KEY", ""),
        "authDomain": os.getenv("FIREBASE_AUTH_DOMAIN", ""),
        "projectId": os.getenv("FIREBASE_PROJECT_ID", ""),
        "storageBucket": os.getenv("FIREBASE_STORAGE_BUCKET", ""),
        "messagingSenderId": os.getenv("FIREBASE_MESSAGING_SENDER_ID", ""),
        "appId": os.getenv("FIREBASE_APP_ID", ""),
        "measurementId": os.getenv("FIREBASE_MEASUREMENT_ID", "")
    }
    return jsonify(config)

@app.route('/cargar', methods=['POST'])
def cargar_archivo():
    if 'archivo' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    archivo = request.files['archivo']
    
    if archivo.filename == '':
        return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
    
    # Aceptar más formatos de Excel
    extensiones_validas = ('.xlsx', '.xlsm', '.xlsb', '.xls')
    nombre_archivo = archivo.filename.lower() if archivo.filename else ''
    
    if not nombre_archivo.endswith(extensiones_validas):
        return jsonify({'error': f'Solo se permiten archivos Excel ({", ".join(extensiones_validas)})'}), 400
    
    try:
        # Procesar el archivo directamente desde memoria
        contenido = archivo.read()
        
        if not contenido or len(contenido) == 0:
            return jsonify({'error': 'El archivo está vacío'}), 400
        
        file_stream = BytesIO(contenido)
        datos_por_mes = extraer_datos_excel(file_stream)
        
        if not datos_por_mes:
            return jsonify({'error': 'No se encontraron datos en las hojas de meses del archivo. Verifica que el archivo tenga hojas con nombres de meses (Enero, Febrero, etc.) y que contengan datos con encabezados (Cedula, Nombres, etc.)'}), 400
        
        datos_por_distrito = organizar_por_distrito(datos_por_mes)
        
        # Ordenar meses según calendario
        meses_ordenados = [m for m in MESES if m in datos_por_mes]
        
        return jsonify({
            'success': True,
            'datos_por_mes': datos_por_mes,
            'datos_por_distrito': datos_por_distrito,
            'meses': meses_ordenados
        })
    except Exception as e:
        print(f'Error al procesar archivo: {traceback.format_exc()}')
        mensaje_error = str(e)
        if 'not a zip file' in mensaje_error.lower() or 'badzip' in mensaje_error.lower():
            mensaje_error = 'El archivo no es un archivo Excel válido o está dañado. Intenta guardarlo de nuevo como .xlsx desde Excel.'
        elif 'permission' in mensaje_error.lower():
            mensaje_error = 'Error de permisos al abrir el archivo. Asegúrate de que no esté abierto en otro programa.'
        return jsonify({'error': f'Error al procesar el archivo: {mensaje_error}'}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
